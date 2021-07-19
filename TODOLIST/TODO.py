import pandas as pd
import datetime
import locale
from datetime import timedelta
import openpyxl
from openpyxl.styles import Font
import requests
import schedule
import time

class Todo:
    def __init__(self):
        locale.setlocale(locale.LC_CTYPE , "Japanese_Japan.932")
        time_date = datetime.datetime.now()
        self.now = time_date.replace(minute=0,second=0,microsecond=0)
        self.tomorrow =  (time_date + timedelta(days=1)).replace(minute=0,second=0,microsecond=0)
        strnow = self.now.strftime('%Y年%m月%d日')
        self.file_name = strnow+'to_do'+'.xlsx'

    def make_excel_schedule(self):
        date_range = pd.date_range(self.now,self.tomorrow,freq='H')
        df_Todo = pd.DataFrame(index=date_range,columns={'やるべきことリスト'})
        df_Todo.to_excel(self.file_name)
        wb = openpyxl.load_workbook(self.file_name)
        ws = wb.worksheets[0]
        ws.title = 'Todo'
        font = Font(name='メイリオ',size=14)
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 40
        for col in ws['A1':'B26']:
            for p in col:
                ws[p.coordinate].font = font
        wb.save(self.file_name)
        wb.close()
    def send_line(self):
        self.df_Todo = pd.read_excel(self.file_name,index_col=0)
        pd_now = pd.to_datetime(self.now)
        for index_num in self.df_Todo.index:
            if pd_now == index_num:
                self.p = self.df_Todo.index.get_loc(pd_now)
                to = self.df_Todo.iloc[self.p,0]
                body = '{0}に、{1}という予定があります'.format(str(self.now), to)
                TOKEN = ''
                line_api = 'https://notify-api.line.me/api/notify'
                TOKEN_dic = {'Authorization': 'Bearer' + ' ' + TOKEN}
                message_dic = {'message': body}
                requests.post(line_api, headers=TOKEN_dic, data=message_dic)
            else:
                pass
def main():
    call_class = Todo()
    time_date =datetime.datetime.now()
    now = time_date.replace(minute=0, second=0, microsecond=0)
    strnow = now.strftime('%Y年%m月%d日')
    file_name = strnow + 'to_do' + '.xlsx'
    try:
        df_Todo_ = pd.read_excel(file_name,index_col=0)
        call_class.send_line()
    except FileNotFoundError:
        call_class.make_excel_schedule()
        df_Todo = pd.read_excel(file_name, index_col=0)
        call_class.send_line()


schedule.every(1).minutes.do(main)
while True:
    schedule.run_pending()
    time.sleep(1)